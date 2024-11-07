from dotenv import load_dotenv
load_dotenv()

from flask import Flask, request, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_migrate import Migrate
from werkzeug.security import generate_password_hash, check_password_hash
import os
import jwt as pyjwt  # Importar con alias para evitar conflictos
from datetime import datetime, timezone, timedelta
from functools import wraps
from io import BytesIO
from sqlalchemy.exc import SQLAlchemyError
import hashlib
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

app = Flask(__name__)

# Configuración de la base de datos
database_url = os.environ.get('DATABASE_URL')
if database_url and database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'fe4d61b2ad570a03abc4910e9d10362f1e3a24ce334d8b22')

db = SQLAlchemy(app)
migrate = Migrate(app, db)

# Configuración de Rate Limiting
limiter = Limiter(
    key_func=get_remote_address,
    default_limits=["200 per day", "50 per hour"]
)

limiter.init_app(app)
# Modelos

class Usuario(db.Model):
    __tablename__ = 'usuario'
    user_id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    apellido = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    rol = db.Column(db.String(50), nullable=False)
    codigo_qr = db.Column(db.String(255), nullable=True)
    fecha_registro = db.Column(db.DateTime, default=datetime.now(timezone.utc))
    ultimo_login = db.Column(db.DateTime)
    asistencias = db.relationship('Asistencia', backref='usuario', lazy=True)
    reportes = db.relationship('Reporte', backref='usuario', lazy=True)
    leaderboard = db.relationship('Leaderboard', backref='usuario', uselist=False)
    aulas = db.relationship('Aula', backref='docente', lazy=True, foreign_keys='Aula.docente_id')

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class DirectorSesion(db.Model):
    __tablename__ = 'director_sesiones'
    sesion_id = db.Column(db.Integer, primary_key=True)
    fecha_sesion = db.Column(db.Date, nullable=False)
    activa = db.Column(db.Boolean, default=True)


# Modelo de Aulas
class Aula(db.Model):
    __tablename__ = 'aulas'
    aula_id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    turno = db.Column(db.String(50), nullable=False)
    docente_id = db.Column(db.Integer, db.ForeignKey('usuario.user_id'), nullable=True)

    asistencias = db.relationship('Asistencia', backref='aula', lazy=True)
    sesiones_qr = db.relationship('SesionQR', backref='aula', lazy=True)

# Modelo de Asistencias
class Asistencia(db.Model):
    __tablename__ = 'asistencias'
    asistencia_id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('usuario.user_id'), nullable=False, index=True)
    aula_id = db.Column(db.Integer, db.ForeignKey('aulas.aula_id'), nullable=False, index=True)
    fecha_asistencia = db.Column(db.Date, nullable=False, index=True)
    hora_entrada = db.Column(db.Time)
    hora_salida = db.Column(db.Time)
    estado = db.Column(db.String(50), nullable=False)

# Modelo de SesionesQR
class SesionQR(db.Model):
    __tablename__ = 'sesionesqr'
    sesion_id = db.Column(db.Integer, primary_key=True)
    aula_id = db.Column(db.Integer, db.ForeignKey('aulas.aula_id'), nullable=False)
    docente_id = db.Column(db.Integer, db.ForeignKey('usuario.user_id'), nullable=False)
    fecha_sesion = db.Column(db.Date, nullable=False)
    hora_inicio = db.Column(db.Time, nullable=False)
    hora_fin = db.Column(db.Time, nullable=False)
    tolerancia_minutos = db.Column(db.Integer, default=0)

# Modelo de Reportes
class Reporte(db.Model):
    __tablename__ = 'reportes'
    reporte_id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('usuario.user_id'), nullable=False)
    descripcion = db.Column(db.Text, nullable=False)
    fecha_reporte = db.Column(db.DateTime, default=datetime.now(timezone.utc))

# Modelo de Leaderboard
class Leaderboard(db.Model):
    __tablename__ = 'leaderboard'
    leaderboard_id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('usuario.user_id'), unique=True, nullable=False)
    puntos = db.Column(db.Integer, default=0, index=True)
    fecha_actualizacion = db.Column(db.DateTime, default=datetime.now(timezone.utc))

# Modelo de Alertas
class Alerta(db.Model):
    __tablename__ = 'alertas'
    alerta_id = db.Column(db.Integer, primary_key=True)
    tipo = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text, nullable=False)
    fecha_creacion = db.Column(db.DateTime, default=datetime.now(timezone.utc))

class AsistenciaDocente(db.Model):
    __tablename__ = 'asistencias_docentes'
    asistencia_id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('usuario.user_id'), nullable=False)
    fecha_asistencia = db.Column(db.Date, nullable=False)
    hora_entrada = db.Column(db.Time)
    hora_salida = db.Column(db.Time)

    usuario = db.relationship('Usuario', backref=db.backref('asistencias_docentes', lazy=True))

# Decorador para rutas protegidas
def token_requerido(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        if 'Authorization' in request.headers:
            try:
                token = request.headers['Authorization'].split(" ")[1]
            except IndexError:
                return jsonify({'mensaje': 'Token mal formado!'}), 401
        if not token:
            return jsonify({'mensaje': 'Token está ausente!'}), 401
        try:
            data = pyjwt.decode(token, app.config['SECRET_KEY'], algorithms=["HS256"])
            current_user = Usuario.query.filter_by(user_id=data['user_id']).first()
            if not current_user:
                return jsonify({'mensaje': 'Usuario no encontrado!'}), 401
        except pyjwt.ExpiredSignatureError:
            return jsonify({'mensaje': 'Token expirado!'}), 401
        except pyjwt.InvalidTokenError:
            return jsonify({'mensaje': 'Token inválido!'}), 401
        except Exception as e:
            app.logger.error(f"Error al decodificar el token: {str(e)}")
            return jsonify({'mensaje': 'Error interno del servidor!'}), 500
        return f(current_user, *args, **kwargs)
    return decorated

# Decorador para verificar roles
def role_required(roles):
    def decorator(f):
        @wraps(f)
        def decorated(current_user, *args, **kwargs):
            if current_user.rol not in roles:
                return jsonify({'mensaje': 'No tienes permiso para realizar esta acción!'}), 403
            return f(current_user, *args, **kwargs)
        return decorated
    return decorator

# Ruta para la raíz '/'
@app.route('/', methods=['GET'])
def index():
    return jsonify({'mensaje': 'Bienvenido a la API de Asistencia QR!'}), 200

# Ruta de registro
# Ruta de registro
@app.route('/registro', methods=['POST'])
def registro():
    try:
        data = request.get_json()
        nombre = data.get('nombre')
        apellido = data.get('apellido')
        email = data.get('email')
        password = data.get('password')

        if not all([nombre, apellido, email, password]):
            return jsonify({'mensaje': 'Faltan datos!'}), 400

        if Usuario.query.filter_by(email=email).first():
            return jsonify({'mensaje': 'El email ya está registrado!'}), 400

        # Generar código QR único
        codigo_qr = generar_codigo_qr_unico_estudiante(email)

        nuevo_usuario = Usuario(
            nombre=nombre,
            apellido=apellido,
            email=email,
            rol='estudiante',
            codigo_qr=codigo_qr  # Guardar el código QR
        )
        nuevo_usuario.set_password(password)
        db.session.add(nuevo_usuario)
        db.session.commit()

        return jsonify({'mensaje': 'Usuario registrado exitosamente!', 'codigo_qr': codigo_qr}), 201
    except Exception as e:
        app.logger.error(f"Error en registro: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

def generar_codigo_qr_unico_estudiante(email):
    """
    Genera un código QR único para un estudiante basado en su email y una clave secreta.
    """
    data = f"{email}-{app.config['SECRET_KEY']}"
    codigo_qr = hashlib.sha256(data.encode()).hexdigest()
    return codigo_qr

def verificar_sesion_activa(sesion_qr, current_datetime):
    current_time = current_datetime.time()
    current_date = current_datetime.date()

    if current_date != sesion_qr.fecha_sesion:
        return {'mensaje': 'La sesión no está activa hoy!'}, 400

    if not (sesion_qr.hora_inicio <= current_time <= sesion_qr.hora_fin):
        return {'mensaje': 'La sesión no está activa en este momento!'}, 400

    return None, None  # No hay error



@app.route('/docente/asistencia/estudiante', methods=['POST'])
@token_requerido
@role_required(['docente'])
def registrar_asistencia_estudiante_por_docente(current_user):
    try:
        data = request.get_json()
        codigo_qr_estudiante = data.get('codigo_qr')
        sesion_id = data.get('sesion_id')

        if not all([codigo_qr_estudiante, sesion_id]):
            return jsonify({'mensaje': 'Faltan datos!'}), 400

        # Verificar que el código QR pertenece a un estudiante
        estudiante = Usuario.query.filter_by(codigo_qr=codigo_qr_estudiante, rol='estudiante').first()
        if not estudiante:
            return jsonify({'mensaje': 'Código QR inválido!'}), 400

        # Obtener la sesión
        sesion_qr = SesionQR.query.get(sesion_id)
        if not sesion_qr:
            return jsonify({'mensaje': 'Sesión no encontrada!'}), 404

        # Verificar que el docente es el asignado al aula
        if sesion_qr.docente_id != current_user.user_id:
            return jsonify({'mensaje': 'No tienes permiso para registrar asistencia en esta sesión!'}), 403

        # Verificar que la sesión está activa
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_datetime = datetime.now(lima_timezone)

        error_response, error_status = verificar_sesion_activa(sesion_qr, current_datetime)
        if error_response:
            return jsonify(error_response), error_status

        # Registrar la asistencia
        response, status_code = registrar_asistencia_estudiante(estudiante.user_id, sesion_qr, current_datetime)
        return jsonify(response), status_code

    except Exception as e:
        app.logger.error(f"Error al registrar asistencia del estudiante por docente: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500


@app.route('/usuarios', methods=['POST'])
@token_requerido
@role_required(['admin'])
def crear_usuario_por_admin(current_user):
    try:
        data = request.get_json()
        nombre = data.get('nombre')
        apellido = data.get('apellido')
        email = data.get('email')
        password = data.get('password')
        rol = data.get('rol')

        if not all([nombre, apellido, email, password, rol]):
            return jsonify({'mensaje': 'Faltan datos!'}), 400

        if rol not in ['docente', 'director']:
            return jsonify({'mensaje': 'Rol inválido! Solo puedes crear docentes o directores.'}), 400

        if Usuario.query.filter_by(email=email).first():
            return jsonify({'mensaje': 'El email ya está registrado!'}), 400

        nuevo_usuario = Usuario(
            nombre=nombre,
            apellido=apellido,
            email=email,
            rol=rol
        )
        nuevo_usuario.set_password(password)
        db.session.add(nuevo_usuario)
        db.session.commit()

        return jsonify({'mensaje': f'Usuario {rol} creado exitosamente!'}), 201
    except Exception as e:
        app.logger.error(f"Error al crear usuario por admin: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500


# Ruta de login
@app.route('/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')

        if not all([email, password]):
            return jsonify({'mensaje': 'Faltan datos!'}), 400

        usuario = Usuario.query.filter_by(email=email).first()
        if not usuario or not usuario.check_password(password):
            return jsonify({'mensaje': 'Credenciales inválidas!'}), 401

        token = pyjwt.encode({
            'user_id': usuario.user_id,
            'exp': datetime.now(timezone.utc) + timedelta(hours=24)
        }, app.config['SECRET_KEY'], algorithm="HS256")

        return jsonify({'token': token}), 200
    except Exception as e:
        app.logger.error(f"Error en login: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

# Ruta protegida
@app.route('/protegido', methods=['GET'])
@token_requerido
def protegido(current_user):
    return jsonify({'mensaje': f'Hola {current_user.nombre}, tienes acceso a esta ruta protegida!'}), 200

# Rutas CRUD para Aulas
@app.route('/aulas', methods=['POST'])
@token_requerido
@role_required(['admin'])
def crear_aula(current_user):
    try:
        data = request.get_json()
        nombre = data.get('nombre')
        turno = data.get('turno')
        docente_id = data.get('docente_id')

        if not all([nombre, turno]):
            return jsonify({'mensaje': 'Faltan datos!'}), 400

        if turno not in ['Mañana', 'Tarde', 'Noche']:
            return jsonify({'mensaje': 'Turno inválido!'}), 400

        nueva_aula = Aula(nombre=nombre, turno=turno, docente_id=docente_id)
        db.session.add(nueva_aula)
        db.session.commit()

        return jsonify({'mensaje': 'Aula creada exitosamente!'}), 201
    except Exception as e:
        app.logger.error(f"Error al crear aula: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

@app.route('/aulas', methods=['GET'])
@token_requerido
def obtener_aulas(current_user):
    try:
        aulas = Aula.query.all()
        resultado = []
        for aula in aulas:
            aula_data = {
                'aula_id': aula.aula_id,
                'nombre': aula.nombre,
                'turno': aula.turno,
                'docente_id': aula.docente_id
            }
            resultado.append(aula_data)
        return jsonify({'aulas': resultado}), 200
    except Exception as e:
        app.logger.error(f"Error al obtener aulas: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

@app.route('/aulas/<int:aula_id>', methods=['GET'])
@token_requerido
def obtener_aula(current_user, aula_id):
    try:
        aula = Aula.query.get(aula_id)
        if not aula:
            return jsonify({'mensaje': 'Aula no encontrada!'}), 404

        aula_data = {
            'aula_id': aula.aula_id,
            'nombre': aula.nombre,
            'turno': aula.turno,
            'docente_id': aula.docente_id
        }
        return jsonify({'aula': aula_data}), 200
    except Exception as e:
        app.logger.error(f"Error al obtener aula: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

@app.route('/aulas/<int:aula_id>', methods=['PUT'])
@token_requerido
@role_required(['admin'])
def actualizar_aula(current_user, aula_id):
    try:
        aula = Aula.query.get(aula_id)
        if not aula:
            return jsonify({'mensaje': 'Aula no encontrada!'}), 404

        data = request.get_json()
        nombre = data.get('nombre')
        turno = data.get('turno')
        docente_id = data.get('docente_id')

        if nombre:
            aula.nombre = nombre
        if turno:
            if turno not in ['Mañana', 'Tarde', 'Noche']:
                return jsonify({'mensaje': 'Turno inválido!'}), 400
            aula.turno = turno
        if docente_id is not None:
            aula.docente_id = docente_id

        db.session.commit()
        return jsonify({'mensaje': 'Aula actualizada exitosamente!'}), 200
    except Exception as e:
        app.logger.error(f"Error al actualizar aula: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

@app.route('/aulas/<int:aula_id>', methods=['DELETE'])
@token_requerido
@role_required(['admin'])
def eliminar_aula(current_user, aula_id):
    try:
        aula = Aula.query.get(aula_id)
        if not aula:
            return jsonify({'mensaje': 'Aula no encontrada!'}), 404

        db.session.delete(aula)
        db.session.commit()
        return jsonify({'mensaje': 'Aula eliminada exitosamente!'}), 200
    except Exception as e:
        app.logger.error(f"Error al eliminar aula: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

@app.route('/director/sesion', methods=['POST'])
@token_requerido
@role_required(['director'])
def crear_sesion_director(current_user):
    try:
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        today = datetime.now(lima_timezone).date()

        # Verificar si ya existe una sesión para hoy
        sesion_existente = DirectorSesion.query.filter_by(fecha_sesion=today, activa=True).first()
        if sesion_existente:
            return jsonify({'mensaje': 'Ya existe una sesión activa para hoy!'}), 400

        nueva_sesion = DirectorSesion(fecha_sesion=today)
        db.session.add(nueva_sesion)
        db.session.commit()

        return jsonify({'mensaje': 'Sesión de director creada exitosamente!', 'sesion_id': nueva_sesion.sesion_id}), 201
    except Exception as e:
        app.logger.error(f"Error al crear sesión de director: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

@app.route('/docente/asistencia', methods=['POST'])
@token_requerido
@role_required(['docente'])
def registrar_asistencia_docente(current_user):
    try:
        data = request.get_json()
        accion = data.get('accion')
        codigo_qr_provided = data.get('codigo_qr')

        if accion not in ['entrada', 'salida']:
            return jsonify({'mensaje': 'Acción inválida!'}), 400

        if not codigo_qr_provided:
            return jsonify({'mensaje': 'Faltan datos! El campo "codigo_qr" es obligatorio.'}), 400

        # Verificar que hay una sesión activa del director
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_datetime = datetime.now(lima_timezone)
        current_date = current_datetime.date()
        current_time = current_datetime.time()

        sesion_director = DirectorSesion.query.filter_by(fecha_sesion=current_date, activa=True).first()
        if not sesion_director:
            return jsonify({'mensaje': 'No hay una sesión activa para hoy!'}), 400

        # Generar el código QR actual para el docente
        codigo_qr_actual = generar_codigo_qr_dinamico_docente(current_user.user_id)

        # Verificar si el código proporcionado coincide
        if codigo_qr_provided != codigo_qr_actual:
            return jsonify({'mensaje': 'Código QR inválido o expirado!'}), 400

        # Verificar si ya existe un registro de asistencia para hoy
        asistencia = AsistenciaDocente.query.filter_by(
            user_id=current_user.user_id,
            fecha_asistencia=current_date
        ).first()

        if not asistencia:
            asistencia = AsistenciaDocente(
                user_id=current_user.user_id,
                fecha_asistencia=current_date
            )
            db.session.add(asistencia)

        if accion == 'entrada':
            if asistencia.hora_entrada:
                return jsonify({'mensaje': 'Ya registraste tu hora de entrada!'}), 400
            asistencia.hora_entrada = current_time
        elif accion == 'salida':
            if asistencia.hora_salida:
                return jsonify({'mensaje': 'Ya registraste tu hora de salida!'}), 400
            asistencia.hora_salida = current_time

        db.session.commit()

        return jsonify({'mensaje': f'Hora de {accion} registrada exitosamente!'}), 201
    except Exception as e:
        app.logger.error(f"Error al registrar asistencia de docente: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500


@app.route('/docente/asistencia/codigo', methods=['GET'])
@token_requerido
@role_required(['docente'])
def obtener_codigo_qr_docente(current_user):
    try:
        # Verificar que hay una sesión activa del director
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_date = datetime.now(lima_timezone).date()

        sesion_director = DirectorSesion.query.filter_by(fecha_sesion=current_date, activa=True).first()
        if not sesion_director:
            return jsonify({'mensaje': 'No hay una sesión activa para hoy!'}), 400

        # Generar el código QR dinámico para el docente
        codigo_qr = generar_codigo_qr_dinamico_docente(current_user.user_id)

        return jsonify({'codigo_qr': codigo_qr}), 200
    except Exception as e:
        app.logger.error(f"Error al obtener código QR para docente: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500


# Ruta para crear Sesión QR
@app.route('/sesionqr', methods=['POST'])
@token_requerido
@role_required(['docente'])
def crear_sesion_qr(current_user):
    try:
        data = request.get_json()
        aula_id = data.get('aula_id')
        fecha_sesion = data.get('fecha_sesion')
        hora_inicio = data.get('hora_inicio')
        hora_fin = data.get('hora_fin')
        tolerancia_minutos = data.get('tolerancia_minutos', 0)

        if not all([aula_id, fecha_sesion, hora_inicio, hora_fin]):
            return jsonify({'mensaje': 'Faltan datos!'}), 400

        # Verificar que el aula existe
        aula = Aula.query.get(aula_id)
        if not aula:
            return jsonify({'mensaje': 'Aula no encontrada!'}), 404

        # Verificar que el aula tiene un docente asignado
        if not aula.docente_id:
            return jsonify({'mensaje': 'El aula no tiene un docente asignado!'}), 400

        # Verificar que el docente actual es el asignado al aula
        if aula.docente_id != current_user.user_id:
            return jsonify({'mensaje': 'No tienes permiso para crear una sesión en esta aula!'}), 403

        # Convertir fechas y horas
        try:
            fecha_sesion_dt = datetime.strptime(fecha_sesion, '%Y-%m-%d').date()
            hora_inicio_tm = datetime.strptime(hora_inicio, '%H:%M:%S').time()
            hora_fin_tm = datetime.strptime(hora_fin, '%H:%M:%S').time()
        except ValueError as e:
            return jsonify({'mensaje': f'Formato de fecha u hora inválido: {str(e)}'}), 400

        # Verificar que hora_inicio es anterior a hora_fin
        if hora_inicio_tm >= hora_fin_tm:
            return jsonify({'mensaje': 'La hora de inicio debe ser anterior a la hora de fin!'}), 400

        nueva_sesion_qr = SesionQR(
            aula_id=aula_id,
            docente_id=current_user.user_id,
            fecha_sesion=fecha_sesion_dt,
            hora_inicio=hora_inicio_tm,
            hora_fin=hora_fin_tm,
            tolerancia_minutos=tolerancia_minutos
        )
        db.session.add(nueva_sesion_qr)
        db.session.commit()

        return jsonify({'mensaje': 'Sesión QR creada exitosamente!', 'sesion_id': nueva_sesion_qr.sesion_id}), 201
    except SQLAlchemyError as e:
        db.session.rollback()
        app.logger.error(f"Error de base de datos al crear sesión QR: {str(e)}")
        return jsonify({'mensaje': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        app.logger.error(f"Error al crear sesión QR: {str(e)}")
        return jsonify({'mensaje': f'Error interno del servidor: {str(e)}'}), 500

def generate_unique_qr_code():
    import uuid
    return str(uuid.uuid4())

def generar_codigo_qr_dinamico(entity_id, entity_type):
    """
    Genera un código QR dinámico basado en el ID de la entidad y su tipo.

    :param entity_id: ID de la entidad (sesion_id para estudiantes, user_id para docentes)
    :param entity_type: Tipo de entidad ('sesion' o 'docente')
    :return: Código QR generado como una cadena hexadecimal
    """
    lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
    current_time = datetime.now(lima_timezone)
    current_minute = current_time.strftime('%Y%m%d%H%M')  # AñoMesDíaHoraMinuto

    # Concatenamos entity_type, entity_id, current_minute y SECRET_KEY para generar el hash
    data = f"{entity_type}-{entity_id}-{current_minute}-{app.config['SECRET_KEY']}"
    codigo_qr = hashlib.sha256(data.encode()).hexdigest()
    return codigo_qr


def generar_codigo_qr_dinamico_estudiante(sesion_id):
    return generar_codigo_qr_dinamico(sesion_id, 'sesion')

def generar_codigo_qr_dinamico_docente(user_id):
    return generar_codigo_qr_dinamico(user_id, 'docente')


@app.route('/sesionqr/<int:sesion_id>/codigo', methods=['GET'])
@token_requerido
@role_required(['docente', 'estudiante'])
def obtener_codigo_qr(current_user, sesion_id):
    try:
        # Obtener la sesión
        sesion = SesionQR.query.get(sesion_id)
        if not sesion:
            return jsonify({'mensaje': 'Sesión no encontrada!'}), 404

        # Verificar que la sesión está activa
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_time = datetime.now(lima_timezone).time()

        if not (sesion.hora_inicio <= current_time <= sesion.hora_fin):
            return jsonify({'mensaje': 'La sesión no está activa!'}), 400

        # Generar el código QR dinámico
        codigo_qr = generar_codigo_qr_dinamico(sesion_id, 'sesion')

        return jsonify({'codigo_qr': codigo_qr}), 200
    except Exception as e:
        app.logger.error(f"Error al obtener código QR: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

# Ruta para registrar asistencia (estudiante)
@app.route('/asistencias', methods=['POST'])
@token_requerido
def registrar_asistencia(current_user):
    if current_user.rol != 'estudiante':
        return jsonify({'mensaje': 'No tienes permiso para realizar esta acción!'}), 403

    try:
        data = request.get_json()
        sesion_id = data.get('sesion_id')
        codigo_qr_provided = data.get('codigo_qr')
        if not all([sesion_id, codigo_qr_provided]):
            return jsonify({'mensaje': 'Faltan datos!'}), 400

        # Obtener la sesión
        sesion_qr = SesionQR.query.get(sesion_id)
        if not sesion_qr:
            return jsonify({'mensaje': 'Sesión no encontrada!'}), 404

        # Verificar que la sesión está activa
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_datetime = datetime.now(lima_timezone)

        error_response, error_status = verificar_sesion_activa(sesion_qr, current_datetime)
        if error_response:
            return jsonify(error_response), error_status

        # Generar el código QR actual para la sesión
        codigo_qr_actual = generar_codigo_qr_dinamico(sesion_id, 'sesion')

        # Verificar si el código proporcionado coincide
        if codigo_qr_provided != codigo_qr_actual:
            return jsonify({'mensaje': 'Código QR inválido o expirado!'}), 400

        # Registrar la asistencia
        response, status_code = registrar_asistencia_estudiante(current_user.user_id, sesion_qr, current_datetime)
        return jsonify(response), status_code

    except Exception as e:
        app.logger.error(f"Error al registrar asistencia: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

def registrar_asistencia_estudiante(estudiante_id, sesion_qr, current_datetime):
    # Verificar si el estudiante ya registró asistencia
    asistencia_existente = Asistencia.query.filter_by(
        user_id=estudiante_id,
        aula_id=sesion_qr.aula_id,
        fecha_asistencia=current_datetime.date()
    ).first()

    if asistencia_existente:
        return {'mensaje': 'El estudiante ya tiene registrada su asistencia para esta sesión!'}, 400

    # Registrar la asistencia
    estado = 'asistió'
    hora_llegada = current_datetime.time()

    # Determinar si el estudiante llegó tarde
    hora_entrada_permitida = (datetime.combine(current_datetime.date(), sesion_qr.hora_inicio) +
                              timedelta(minutes=sesion_qr.tolerancia_minutos)).time()

    if hora_llegada > hora_entrada_permitida:
        estado = 'tardanza'

    nueva_asistencia = Asistencia(
        user_id=estudiante_id,
        aula_id=sesion_qr.aula_id,
        fecha_asistencia=current_datetime.date(),
        hora_entrada=hora_llegada,
        estado=estado
    )
    db.session.add(nueva_asistencia)
    db.session.commit()

    return {'mensaje': 'Asistencia registrada exitosamente!', 'estado': estado}, 201


# Ruta para obtener asistencias con filtros avanzados
@app.route('/asistencias', methods=['GET'])
@token_requerido
def obtener_asistencias(current_user):
    try:
        # Obtener parámetros de filtro
        aula_id = request.args.get('aula_id', type=int)
        docente_id = request.args.get('docente_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        query = Asistencia.query

        if aula_id:
            query = query.filter_by(aula_id=aula_id)

        if docente_id:
            # Filtrar asistencias por aulas del docente
            query = query.join(Aula).filter(Aula.docente_id == docente_id)

        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, '%d/%m/%Y').date()
            query = query.filter(Asistencia.fecha_asistencia >= fecha_inicio)

        if fecha_fin:
            fecha_fin = datetime.strptime(fecha_fin, '%d/%m/%Y').date()
            query = query.filter(Asistencia.fecha_asistencia <= fecha_fin)

        asistencias = query.all()
        resultado = []
        for asistencia in asistencias:
            asistencia_data = {
                'asistencia_id': asistencia.asistencia_id,
                'user_id': asistencia.user_id,
                'aula_id': asistencia.aula_id,
                'fecha_asistencia': asistencia.fecha_asistencia.strftime('%d/%m/%Y'),
                'hora_entrada': asistencia.hora_entrada.strftime('%H:%M:%S') if asistencia.hora_entrada else None,
                'hora_salida': asistencia.hora_salida.strftime('%H:%M:%S') if asistencia.hora_salida else None,
                'estado': asistencia.estado
            }
            resultado.append(asistencia_data)
        return jsonify({'asistencias': resultado}), 200
    except Exception as e:
        app.logger.error(f"Error al obtener asistencias: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

# Ruta para obtener el leaderboard
@app.route('/leaderboard', methods=['GET'])
@token_requerido
def obtener_leaderboard(current_user):
    try:
        leaderboard = Leaderboard.query.join(Usuario).order_by(Leaderboard.puntos.desc()).all()
        resultado = []
        for entry in leaderboard:
            usuario = entry.usuario
            entry_data = {
                'user_id': entry.user_id,
                'nombre': usuario.nombre,
                'apellido': usuario.apellido,
                'puntos': entry.puntos,
                'fecha_actualizacion': entry.fecha_actualizacion.strftime('%d/%m/%Y %H:%M:%S')
            }
            resultado.append(entry_data)
        return jsonify({'leaderboard': resultado}), 200
    except Exception as e:
        app.logger.error(f"Error al obtener leaderboard: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

# Ruta para generar alertas
@app.route('/asistencias/exportar', methods=['GET'])
@token_requerido
@role_required(['docente', 'admin', 'director'])
def exportar_asistencias(current_user):
    try:
        # Obtener parámetros
        opcion = request.args.get('opcion')  # 'semanal', 'mensual', 'todo'
        aula_id = request.args.get('aula_id', type=int)

        if not all([opcion, aula_id]):
            return jsonify({'mensaje': 'Faltan datos! Debes proporcionar la opción y aula_id.'}), 400

        if opcion not in ['semanal', 'mensual', 'todo']:
            return jsonify({'mensaje': 'Opción inválida! Debe ser "semanal", "mensual" o "todo".'}), 400

        # Obtener la fecha actual en la zona horaria de Lima
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_date = datetime.now(lima_timezone).date()

        # Inicializar fecha_inicio
        fecha_inicio = None

        # Calcular el rango de fechas según la opción
        if opcion == 'semanal':
            fecha_inicio = current_date - timedelta(days=7)
        elif opcion == 'mensual':
            fecha_inicio = current_date - timedelta(days=30)
        elif opcion == 'todo':
            # Fecha muy antigua para incluir todas las asistencias
            fecha_inicio = datetime.strptime('2000-01-01', '%Y-%m-%d').date()
        else:
            # Este else no debería ser alcanzado, pero lo agregamos para evitar posibles alertas
            return jsonify({'mensaje': 'Opción inválida!'}), 400

        fecha_fin = current_date

        # Verificar permisos y filtrar asistencias según el rol
        if current_user.rol == 'docente':
            # El docente solo puede exportar asistencias de sus propias aulas
            aula = Aula.query.get(aula_id)
            if not aula:
                return jsonify({'mensaje': 'Aula no encontrada!'}), 404
            if aula.docente_id != current_user.user_id:
                return jsonify({'mensaje': 'No tienes permiso para exportar asistencias de esta aula!'}), 403
            # Filtrar asistencias del aula y rango de fechas
            asistencias = Asistencia.query.filter(
                Asistencia.aula_id == aula_id,
                Asistencia.fecha_asistencia >= fecha_inicio,
                Asistencia.fecha_asistencia <= fecha_fin
            ).join(Usuario).order_by(Usuario.apellido, Usuario.nombre).all()
        elif current_user.rol == 'director':
            # El director puede exportar asistencias del aula en el rango de fechas
            asistencias = Asistencia.query.filter(
                Asistencia.aula_id == aula_id,
                Asistencia.fecha_asistencia >= fecha_inicio,
                Asistencia.fecha_asistencia <= fecha_fin
            ).join(Usuario).order_by(Usuario.apellido, Usuario.nombre).all()
        elif current_user.rol == 'admin':
            # El administrador puede exportar todas las asistencias
            asistencias = Asistencia.query.filter(
                Asistencia.aula_id == aula_id,
                Asistencia.fecha_asistencia >= fecha_inicio,
                Asistencia.fecha_asistencia <= fecha_fin
            ).join(Usuario).order_by(Usuario.apellido, Usuario.nombre).all()
        else:
            return jsonify({'mensaje': 'No tienes permiso para exportar asistencias!'}), 403

        if not asistencias:
            return jsonify({'mensaje': 'No hay asistencias registradas en el período especificado.'}), 404

        # Crear una lista de fechas en el rango
        delta = fecha_fin - fecha_inicio
        lista_fechas = [fecha_inicio + timedelta(days=i) for i in range(delta.days + 1)]

        # Crear un diccionario para mapear user_id a datos del usuario
        usuarios = {}
        for asistencia in asistencias:
            user_id = asistencia.user_id
            if user_id not in usuarios:
                usuarios[user_id] = {
                    'apellido': asistencia.usuario.apellido,
                    'nombre': asistencia.usuario.nombre,
                    'asistencias': {}
                }
            usuarios[user_id]['asistencias'][asistencia.fecha_asistencia] = asistencia

        # Obtener todos los usuarios inscritos en el aula (opcional)
        lista_usuarios = list(usuarios.values())

        # Crear el Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencias"

        # Escribir cabeceras
        headers = ['APELLIDOS', 'NOMBRES']
        # Agregar columnas para cada fecha
        for fecha in lista_fechas:
            # Obtener el nombre del día en español
            nombre_dia_es = obtener_nombre_dia_espanol(fecha)
            encabezado = f"{nombre_dia_es} {fecha.strftime('%d/%m/%Y')}"
            headers.append(encabezado)
        ws.append(headers)

        # Escribir datos de asistencias
        for usuario in lista_usuarios:
            row = [usuario['apellido'], usuario['nombre']]
            for fecha in lista_fechas:
                asistencia = usuario['asistencias'].get(fecha)
                if asistencia:
                    estado = asistencia.estado
                    # Concatenar estado y horas si es necesario
                    hora_entrada = asistencia.hora_entrada.strftime('%H:%M:%S') if asistencia.hora_entrada else ''
                    hora_salida = asistencia.hora_salida.strftime('%H:%M:%S') if asistencia.hora_salida else ''
                    contenido_celda = f"{estado}"
                    if current_user.rol == 'director':
                        # Si el usuario es director, incluir horas de entrada y salida
                        contenido_celda += f"\nEntrada: {hora_entrada}\nSalida: {hora_salida}"
                else:
                    contenido_celda = ''
                row.append(contenido_celda)
            ws.append(row)

        # Ajustar el ancho de las columnas y alineación
        for i, column_cells in enumerate(ws.columns, 1):
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(i)].width = length + 5  # Añadir un poco más de ancho
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"asistencias_{opcion}_{aula_id}.xlsx"

        # Enviar el archivo con los headers correctos
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        app.logger.error(f"Error al exportar asistencias: {str(e)}")
        return jsonify({'mensaje': f'Error interno del servidor: {str(e)}'}), 500

def obtener_nombre_dia_espanol(fecha):
    dias_semana = {
        0: 'LUNES',
        1: 'MARTES',
        2: 'MIÉRCOLES',
        3: 'JUEVES',
        4: 'VIERNES',
        5: 'SÁBADO',
        6: 'DOMINGO'
    }
    dia_semana = fecha.weekday()  # Devuelve un número de 0 (lunes) a 6 (domingo)
    return dias_semana.get(dia_semana, '')

if __name__ == '__main__':
    app.run(debug=True)
