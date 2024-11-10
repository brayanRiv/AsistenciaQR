from flask import Blueprint, request, jsonify, send_file, render_template, abort
from app.models import Asistencia, AsistenciaDocente, Leaderboard, Usuario, Aula, SesionQR, DirectorSesion
from app.utils.decorators import token_requerido, role_required
from app.utils.helpers import verificar_sesion_activa, registrar_asistencia_estudiante, obtener_nombre_dia_espanol, \
     generar_codigo_qr_dinamico_docente, generar_codigo_qr_dinamico
from flask import current_app
from app import db
from datetime import datetime, timezone, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

asistencia_bp = Blueprint('asistencia_bp', __name__)

@asistencia_bp.route('/asistencias', methods=['POST'])
@token_requerido
def registrar_asistencia(current_user):
    if current_user.rol != 'estudiante':
        return jsonify({'mensaje': 'No tienes permiso para realizar esta acción!'}), 403

    try:
        data = request.get_json()
        sesion_id = data.get('sesion_id')
        codigo_qr_provided = data.get('codigo_qr')
        if not all([sesion_id, codigo_qr_provided]):
            return jsonify({'mensaje': 'Faltan datos!'}), 400

        # Obtener la sesión
        sesion_qr = SesionQR.query.get(sesion_id)
        if not sesion_qr:
            return jsonify({'mensaje': 'Sesión no encontrada!'}), 404

        # Verificar que la sesión está activa
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_datetime = datetime.now(lima_timezone)

        error_response, error_status = verificar_sesion_activa(sesion_qr, current_datetime)
        if error_response:
            return jsonify(error_response), error_status

        # Generar el código QR actual para la sesión
        codigo_qr_actual = generar_codigo_qr_dinamico(sesion_id, 'sesion')

        # Verificar si el código proporcionado coincide
        if codigo_qr_provided != codigo_qr_actual:
            return jsonify({'mensaje': 'Código QR inválido o expirado!'}), 400

        # Registrar la asistencia
        response, status_code = registrar_asistencia_estudiante(current_user.user_id, sesion_qr, current_datetime)
        return jsonify(response), status_code

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al registrar asistencia: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500


@asistencia_bp.route('/asistencias', methods=['GET'])
@token_requerido
def obtener_asistencias(current_user):
    try:
        # Obtener parámetros de filtro
        aula_id = request.args.get('aula_id', type=int)
        docente_id = request.args.get('docente_id', type=int)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')

        query = Asistencia.query

        if aula_id:
            query = query.filter_by(aula_id=aula_id)

        if docente_id:
            # Filtrar asistencias por aulas del docente
            query = query.join(Aula).filter(Aula.docente_id == docente_id)

        if fecha_inicio:
            fecha_inicio = datetime.strptime(fecha_inicio, '%d/%m/%Y').date()
            query = query.filter(Asistencia.fecha_asistencia >= fecha_inicio)

        if fecha_fin:
            fecha_fin = datetime.strptime(fecha_fin, '%d/%m/%Y').date()
            query = query.filter(Asistencia.fecha_asistencia <= fecha_fin)

        asistencias = query.all()
        resultado = []
        for asistencia in asistencias:
            asistencia_data = {
                'asistencia_id': asistencia.asistencia_id,
                'user_id': asistencia.user_id,
                'aula_id': asistencia.aula_id,
                'fecha_asistencia': asistencia.fecha_asistencia.strftime('%d/%m/%Y'),
                'hora_entrada': asistencia.hora_entrada.strftime('%H:%M:%S') if asistencia.hora_entrada else None,
                'hora_salida': asistencia.hora_salida.strftime('%H:%M:%S') if asistencia.hora_salida else None,
                'estado': asistencia.estado
            }
            resultado.append(asistencia_data)
        return jsonify({'asistencias': resultado}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al obtener asistencias: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

@asistencia_bp.route('/docente/asistencia', methods=['POST'])
@token_requerido
@role_required(['docente'])
def registrar_asistencia_docente(current_user):
    try:
        data = request.get_json()
        accion = data.get('accion')
        codigo_qr_provided = data.get('codigo_qr')

        if accion not in ['entrada', 'salida']:
            return jsonify({'mensaje': 'Acción inválida!'}), 400

        if not codigo_qr_provided:
            return jsonify({'mensaje': 'Faltan datos! El campo "codigo_qr" es obligatorio.'}), 400

        # Verificar que hay una sesión activa del director
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_datetime = datetime.now(lima_timezone)
        current_date = current_datetime.date()
        current_time = current_datetime.time()

        sesion_director = DirectorSesion.query.filter_by(fecha_sesion=current_date, activa=True).first()
        if not sesion_director:
            return jsonify({'mensaje': 'No hay una sesión activa para hoy!'}), 400

        # Generar el código QR actual para el docente
        codigo_qr_actual = generar_codigo_qr_dinamico_docente(current_user.user_id)

        # Verificar si el código proporcionado coincide
        if codigo_qr_provided != codigo_qr_actual:
            return jsonify({'mensaje': 'Código QR inválido o expirado!'}), 400

        # Verificar si ya existe un registro de asistencia para hoy
        asistencia = AsistenciaDocente.query.filter_by(
            user_id=current_user.user_id,
            fecha_asistencia=current_date
        ).first()

        if not asistencia:
            asistencia = AsistenciaDocente(
                user_id=current_user.user_id,
                fecha_asistencia=current_date
            )
            db.session.add(asistencia)

        if accion == 'entrada':
            if asistencia.hora_entrada:
                return jsonify({'mensaje': 'Ya registraste tu hora de entrada!'}), 400
            asistencia.hora_entrada = current_time
        elif accion == 'salida':
            if asistencia.hora_salida:
                return jsonify({'mensaje': 'Ya registraste tu hora de salida!'}), 400
            asistencia.hora_salida = current_time

        db.session.commit()

        return jsonify({'mensaje': f'Hora de {accion} registrada exitosamente!'}), 201
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al registrar asistencia de docente: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

@asistencia_bp.route('/leaderboard', methods=['GET'])
@token_requerido
def obtener_leaderboard(current_user):
    try:
        leaderboard = Leaderboard.query.join(Usuario).order_by(Leaderboard.puntos.desc()).all()
        resultado = []
        for entry in leaderboard:
            usuario = entry.usuario
            entry_data = {
                'user_id': entry.user_id,
                'nombre': usuario.nombre,
                'apellido': usuario.apellido,
                'puntos': entry.puntos,
                'fecha_actualizacion': entry.fecha_actualizacion.strftime('%d/%m/%Y %H:%M:%S')
            }
            resultado.append(entry_data)
        return jsonify({'leaderboard': resultado}), 200
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al obtener leaderboard: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500

# Ruta para generar alertas
@asistencia_bp.route('/asistencias/exportar', methods=['GET'])
@token_requerido
@role_required(['docente', 'admin', 'director'])
def exportar_asistencias(current_user):
    try:
        # Obtener parámetros
        opcion = request.args.get('opcion')  # 'semanal', 'mensual', 'todo'
        aula_id = request.args.get('aula_id', type=int)

        if not all([opcion, aula_id]):
            return jsonify({'mensaje': 'Faltan datos! Debes proporcionar la opción y aula_id.'}), 400

        if opcion not in ['semanal', 'mensual', 'todo']:
            return jsonify({'mensaje': 'Opción inválida! Debe ser "semanal", "mensual" o "todo".'}), 400

        # Obtener la fecha actual en la zona horaria de Lima
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_date = datetime.now(lima_timezone).date()

        # Inicializar fecha_inicio
        fecha_inicio = None

        # Calcular el rango de fechas según la opción
        if opcion == 'semanal':
            fecha_inicio = current_date - timedelta(days=7)
        elif opcion == 'mensual':
            fecha_inicio = current_date - timedelta(days=30)
        elif opcion == 'todo':
            # Fecha muy antigua para incluir todas las asistencias
            fecha_inicio = datetime.strptime('2000-01-01', '%Y-%m-%d').date()
        else:
            # Este else no debería ser alcanzado, pero lo agregamos para evitar posibles alertas
            return jsonify({'mensaje': 'Opción inválida!'}), 400

        fecha_fin = current_date

        # Verificar permisos y filtrar asistencias según el rol
        if current_user.rol == 'docente':
            # El docente solo puede exportar asistencias de sus propias aulas
            aula = Aula.query.get(aula_id)
            if not aula:
                return jsonify({'mensaje': 'Aula no encontrada!'}), 404
            if aula.docente_id != current_user.user_id:
                return jsonify({'mensaje': 'No tienes permiso para exportar asistencias de esta aula!'}), 403
            # Filtrar asistencias del aula y rango de fechas
            asistencias = Asistencia.query.filter(
                Asistencia.aula_id == aula_id,
                Asistencia.fecha_asistencia >= fecha_inicio,
                Asistencia.fecha_asistencia <= fecha_fin
            ).join(Usuario).order_by(Usuario.apellido, Usuario.nombre).all()
        elif current_user.rol == 'director':
            # El director puede exportar asistencias del aula en el rango de fechas
            asistencias = Asistencia.query.filter(
                Asistencia.aula_id == aula_id,
                Asistencia.fecha_asistencia >= fecha_inicio,
                Asistencia.fecha_asistencia <= fecha_fin
            ).join(Usuario).order_by(Usuario.apellido, Usuario.nombre).all()
        elif current_user.rol == 'admin':
            # El administrador puede exportar todas las asistencias
            asistencias = Asistencia.query.filter(
                Asistencia.aula_id == aula_id,
                Asistencia.fecha_asistencia >= fecha_inicio,
                Asistencia.fecha_asistencia <= fecha_fin
            ).join(Usuario).order_by(Usuario.apellido, Usuario.nombre).all()
        else:
            return jsonify({'mensaje': 'No tienes permiso para exportar asistencias!'}), 403

        if not asistencias:
            return jsonify({'mensaje': 'No hay asistencias registradas en el período especificado.'}), 404

        # Crear una lista de fechas en el rango
        delta = fecha_fin - fecha_inicio
        lista_fechas = [fecha_inicio + timedelta(days=i) for i in range(delta.days + 1)]

        # Crear un diccionario para mapear user_id a datos del usuario
        usuarios = {}
        for asistencia in asistencias:
            user_id = asistencia.user_id
            if user_id not in usuarios:
                usuarios[user_id] = {
                    'apellido': asistencia.usuario.apellido,
                    'nombre': asistencia.usuario.nombre,
                    'asistencias': {}
                }
            usuarios[user_id]['asistencias'][asistencia.fecha_asistencia] = asistencia

        # Obtener todos los usuarios inscritos en el aula (opcional)
        lista_usuarios = list(usuarios.values())

        # Crear el Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Asistencias"

        # Escribir cabeceras
        headers = ['APELLIDOS', 'NOMBRES']
        # Agregar columnas para cada fecha
        for fecha in lista_fechas:
            # Obtener el nombre del día en español
            nombre_dia_es = obtener_nombre_dia_espanol(fecha)
            encabezado = f"{nombre_dia_es} {fecha.strftime('%d/%m/%Y')}"
            headers.append(encabezado)
        ws.append(headers)

        # Escribir datos de asistencias
        for usuario in lista_usuarios:
            row = [usuario['apellido'], usuario['nombre']]
            for fecha in lista_fechas:
                asistencia = usuario['asistencias'].get(fecha)
                if asistencia:
                    estado = asistencia.estado
                    # Concatenar estado y horas si es necesario
                    hora_entrada = asistencia.hora_entrada.strftime('%H:%M:%S') if asistencia.hora_entrada else ''
                    hora_salida = asistencia.hora_salida.strftime('%H:%M:%S') if asistencia.hora_salida else ''
                    contenido_celda = f"{estado}"
                    if current_user.rol == 'director':
                        # Si el usuario es director, incluir horas de entrada y salida
                        contenido_celda += f"\nEntrada: {hora_entrada}\nSalida: {hora_salida}"
                else:
                    contenido_celda = ''
                row.append(contenido_celda)
            ws.append(row)

        # Ajustar el ancho de las columnas y alineación
        for i, column_cells in enumerate(ws.columns, 1):
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(i)].width = length + 5  # Añadir un poco más de ancho
            for cell in column_cells:
                cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

        # Guardar en BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"asistencias_{opcion}_{aula_id}.xlsx"

        # Enviar el archivo con los headers correctos
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al crear sesión QR: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor.'}), 500

@asistencia_bp.route('/docente/asistencia/estudiante', methods=['POST'])
@token_requerido
@role_required(['docente'])
def registrar_asistencia_estudiante_por_docente(current_user):
    try:
        data = request.get_json()
        codigo_qr_estudiante = data.get('codigo_qr')
        sesion_id = data.get('sesion_id')

        if not all([codigo_qr_estudiante, sesion_id]):
            return jsonify({'mensaje': 'Faltan datos!'}), 400

        # Verificar que el código QR pertenece a un estudiante
        estudiante = Usuario.query.filter_by(codigo_qr=codigo_qr_estudiante, rol='estudiante').first()
        if not estudiante:
            return jsonify({'mensaje': 'Código QR inválido!'}), 400

        # Obtener la sesión
        sesion_qr = SesionQR.query.get(sesion_id)
        if not sesion_qr:
            return jsonify({'mensaje': 'Sesión no encontrada!'}), 404

        # Verificar que el docente es el asignado al aula
        if sesion_qr.docente_id != current_user.user_id:
            return jsonify({'mensaje': 'No tienes permiso para registrar asistencia en esta sesión!'}), 403

        # Verificar que la sesión está activa
        lima_timezone = timezone(timedelta(hours=-5))  # UTC-5
        current_datetime = datetime.now(lima_timezone)

        error_response, error_status = verificar_sesion_activa(sesion_qr, current_datetime)
        if error_response:
            return jsonify(error_response), error_status

        # Registrar la asistencia
        response, status_code = registrar_asistencia_estudiante(estudiante.user_id, sesion_qr, current_datetime)
        return jsonify(response), status_code

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error al registrar asistencia del estudiante por docente: {str(e)}")
        return jsonify({'mensaje': 'Error interno del servidor!'}), 500
